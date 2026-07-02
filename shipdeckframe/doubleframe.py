# __ IMPORTS __________________________
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from scipy.linalg import eigh
import utlities as ut
import openpyxl as op

# __ GEOMETRY _________________________
    #          3           4                       
    #   (4)---------(5)---------(6)              
    #    |           |           |              _____w_____
    #   5|          6|          7|              |         |
    #    |           |           |              |         | t
    #   (1)---------(2)---------(3)             |_________|
    #          1           2

# __ PROPERTIES _______________________
t = 0.04                #m,     thickness of the frame
w = 0.08                #m,     width of the frame
A = t * w               #m^2,   area of cross-section
I = (w * t**3) / 12     #m^4,   second moment of area
E = 210e+9              #Pa,    Young's Modulus
rho = 7850              #kg/m^3, Density

# __ COORDINATES ______________________
L1 = 4; N1 = 20; dl1 = L1/N1
L2 = 4; N2 = 20; dl2 = L2/N2
L3 = 4; N3 = 20; dl3 = L3/N3
L4 = 4; N4 = 20; dl4 = L4/N4
L5 = 3; N5 = 15; dl5 = L5/N5
L6 = 3; N6 = 15; dl6 = L6/N6
L7 = 3; N7 = 15; dl7 = L7/N7
# __ Beam1 __:
x1 = np.linspace(0, L1, N1+1)
y1 = np.zeros_like(x1)
xy1 = np.array([x1, y1]).T
# __ Beam2 __:
x2 = np.linspace(0, L2, N2+1) + L1
y2 = np.zeros_like(x2)
xy2 = np.array([x2, y2]).T
# __ Beam3 __:
x3 = np.linspace(0, L3, N3+1)
y3 = np.zeros_like(x3) + L5
xy3 = np.array([x3, y3]).T
# __ Beam4 __:
x4 = np.linspace(0, L4, N4+1) + L3
y4 = np.zeros_like(x4) + L5
xy4 = np.array([x4, y4]).T
# __ Beam5 __:
y5 = np.linspace(0, L5, N5+1)
x5 = np.zeros_like(y5) 
xy5 = np.array([x5, y5]).T
# __ Beam6 __:
y6 = np.linspace(0, L6, N6+1)
x6 = np.zeros_like(y6) + L1
xy6 = np.array([x6, y6]).T
# __ Beam7 __:
y7 = np.linspace(0, L7, N7+1)
x7 = np.zeros_like(y7) + L2 + L1
xy7 = np.array([x7, y7]).T
x = np.concatenate([x1, x2, x3, x4, x5, x6, x7])
y = np.concatenate([y1, y2, y3, y4, y5, y6, y7])
# __ Global Nodes __:
xy = np.array([x, y]).T
_, idx = np.unique(xy, axis = 0, return_index = True)
xy = xy[np.sort(idx)]
# __ Node separation __:
nodes = {tuple(n) : i for i, n in enumerate(xy)}
nodes1 = np.array([nodes[tuple(i)] for i in xy1])
nodes2 = np.array([nodes[tuple(i)] for i in xy2])
nodes3 = np.array([nodes[tuple(i)] for i in xy3])
nodes4 = np.array([nodes[tuple(i)] for i in xy4])
nodes5 = np.array([nodes[tuple(i)] for i in xy5])
nodes6 = np.array([nodes[tuple(i)] for i in xy6])
nodes7 = np.array([nodes[tuple(i)] for i in xy7])
details = {1: [nodes1, 0, dl1, x1, y1], 
           2: [nodes2, 0, dl2, x2, y2], 
           3: [nodes3, 0, dl3, x3, y3], 
           4: [nodes4, 0, dl4, x4, y4], 
           5: [nodes5, 90, dl5, x5, y5], 
           6: [nodes6, 90, dl6, x6, y6], 
           7: [nodes7, 90, dl7, x7, y7]}
N = len(xy)  

# __ MATRIX ASSEMBLY __________________
M = np.zeros((3*xy.shape[0], 3*xy.shape[0]))
K = np.zeros_like(M)
for i in details:
    nownodes = details[i][0]
    theta = details[i][1]
    dl = details[i][2]
    T = ut.transformation(theta)
    k = ut.stiffness(T, dl, E, A, I)
    m = ut.mass(T, dl, A, rho)
    for j in range(len(nownodes) - 1):
        ut.matrixform(m, M, j, nownodes)
        ut.matrixform(k, K, j, nownodes)

# __ BOUNDARY CONDITIONS ______________
fixed = np.array([0, 1, 2, 120, 121])
free = np.setdiff1d(np.arange(0, 3*N), fixed)
alldf = np.arange(0, 3*N)
K_red = K[np.ix_(free, free)]
M_red = M[np.ix_(free, free)]

# __ EXPORTING DETAILS ________________
dfdetails = pd.DataFrame({
    'x' : xy[:, 0],
    'y' : xy[:, 1],
    'u' : np.arange(0, 3*N, 3),
    'v' : np.arange(1, 3*N, 3),
    'theta' : np.arange(2, 3*N, 3),
})
dfk = pd.DataFrame(K)
dfm = pd.DataFrame(M)

file = 'shipdeckframe/doubleframe.xlsx'
sheet1 = 'Node_Details'
sheet2 = 'Stiffness'
sheet3 = 'Mass'
sheetnames = {sheet1: dfdetails, sheet2: dfk, sheet3: dfm}
sheets = op.load_workbook(file).sheetnames
with pd.ExcelWriter(file, mode = 'a', engine = 'openpyxl') as writer:
    for sheet in sheetnames:
        if sheet not in sheets:            
            sheetnames[sheet].to_excel(writer, sheet_name=sheet)